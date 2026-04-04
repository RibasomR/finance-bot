"""
Сервис для экспорта данных пользователя в Excel (XLSX).

Содержит функции для создания Excel файлов с транзакциями.
"""

import os
from datetime import datetime, timezone
from typing import Optional
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from loguru import logger

from bot.models import Transaction, TransactionType
from bot.services.database import (
    get_user_transactions_with_filters,
    get_user_statistics,
    get_or_create_user,
)


## Генерация Excel файла с транзакциями
async def generate_transactions_excel(
    user_id: int,
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
) -> str:
    """
    Генерировать Excel файл с транзакциями пользователя.
    
    Создает форматированный XLSX файл со всеми транзакциями пользователя
    за указанный период. Включает цветовое выделение, форматирование
    и итоговую статистику.
    
    :param user_id: ID пользователя
    :param start_date: Начало периода (если None - с начала времени)
    :param end_date: Конец периода (если None - до текущего момента)
    :return: Путь к созданному файлу
    :raises Exception: При ошибке создания файла
    
    Example:
        >>> file_path = await generate_transactions_excel(
        ...     user_id=1,
        ...     start_date=datetime(2024, 1, 1),
        ...     end_date=datetime(2024, 12, 31)
        ... )
    """
    logger.info(f"Начинаю генерацию Excel для пользователя {user_id}")
    
    # Получаем все транзакции
    transactions = await get_user_transactions_with_filters(
        user_id=user_id,
        start_date=start_date,
        end_date=end_date,
        limit=10000,
        offset=0
    )
    
    # Получаем статистику
    stats = await get_user_statistics(user_id, start_date, end_date)
    
    # Создаем Excel файл
    wb = Workbook()
    ws = wb.active
    ws.title = "Транзакции"
    
    # Определяем период для заголовка
    if start_date and end_date:
        period_text = f"с {start_date.strftime('%d.%m.%Y')} по {end_date.strftime('%d.%m.%Y')}"
    elif start_date:
        period_text = f"с {start_date.strftime('%d.%m.%Y')}"
    elif end_date:
        period_text = f"по {end_date.strftime('%d.%m.%Y')}"
    else:
        period_text = "за всё время"
    
    # Стили
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    income_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    expense_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Заголовок документа
    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.value = f"Отчет по транзакциям {period_text}"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")
    
    # Статистика - вычисляем из данных по валютам
    total_income = sum(stats.get('income_by_currency', {}).values(), Decimal('0'))
    total_expense = sum(stats.get('expense_by_currency', {}).values(), Decimal('0'))
    balance = total_income - total_expense

    ws.merge_cells('A2:B2')
    ws['A2'] = "Статистика"
    ws['A2'].font = Font(bold=True, size=12)

    ws['A3'] = "Общий баланс:"
    ws['B3'] = float(balance)
    ws['B3'].number_format = '#,##0.00 ₽'

    ws['A4'] = "Доходы:"
    ws['B4'] = float(total_income)
    ws['B4'].number_format = '#,##0.00 ₽'
    ws['B4'].fill = income_fill

    ws['A5'] = "Расходы:"
    ws['B5'] = float(total_expense)
    ws['B5'].number_format = '#,##0.00 ₽'
    ws['B5'].fill = expense_fill

    ws['A6'] = "Количество операций:"
    ws['B6'] = stats.get('income_count', 0) + stats.get('expense_count', 0)
    
    # Заголовок таблицы транзакций
    headers = ['Дата', 'Время', 'Тип', 'Сумма (₽)', 'Категория', 'Описание']
    header_row = 8
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Заполняем транзакции
    current_row = header_row + 1
    for transaction in transactions:
        # Дата
        ws.cell(row=current_row, column=1, value=transaction.created_at.strftime('%d.%m.%Y'))
        
        # Время
        ws.cell(row=current_row, column=2, value=transaction.created_at.strftime('%H:%M:%S'))
        
        # Тип
        type_cell = ws.cell(
            row=current_row,
            column=3,
            value="Доход" if transaction.type == TransactionType.INCOME else "Расход"
        )
        
        # Сумма
        amount_value = float(transaction.amount)
        if transaction.type == TransactionType.EXPENSE:
            amount_value = -amount_value
        
        amount_cell = ws.cell(row=current_row, column=4, value=amount_value)
        amount_cell.number_format = '#,##0.00'
        
        # Применяем цвет к строке
        fill = income_fill if transaction.type == TransactionType.INCOME else expense_fill
        for col in range(1, 7):
            ws.cell(row=current_row, column=col).fill = fill
            ws.cell(row=current_row, column=col).border = border
        
        # Категория
        ws.cell(
            row=current_row,
            column=5,
            value=f"{transaction.category.emoji} {transaction.category.name}"
        )
        
        # Описание
        ws.cell(
            row=current_row,
            column=6,
            value=transaction.description or ""
        )
        
        current_row += 1
    
    # Настройка ширины колонок
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 40
    
    # Сохраняем файл
    os.makedirs("logs", exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"logs/transactions_export_{user_id}_{timestamp}.xlsx"
    
    wb.save(filename)
    logger.success(f"✅ Excel файл создан: {filename}")
    
    return filename


## Удаление временного файла экспорта
def cleanup_export_file(file_path: str) -> None:
    """
    Удалить временный файл экспорта.
    
    :param file_path: Путь к файлу для удаления
    :return: None
    
    Example:
        >>> cleanup_export_file("logs/transactions_export_1_20240119.xlsx")
    """
    try:
        if os.path.exists(file_path):
            os.remove(file_path)
            logger.info(f"🗑 Удален временный файл: {file_path}")
    except Exception as e:
        logger.error(f"❌ Ошибка удаления файла {file_path}: {e}")

